#!/usr/bin/env python3
"""Build per-annotator Excel workbooks with dropdown validation."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation

CSV_FIELDNAMES = [
    "instance_id",
    "dialogue_id",
    "turn_id",
    "system_context",
    "user_utterance",
    "label",
    "annotator_id",
    "annotation_pass",
    "notes",
]

WORKBOOK_FIELDNAMES = [
    "instance_id",
    "dialogue_id",
    "turn_id",
    "system_context",
    "user_utterance",
    "label",
    "annotator_id",
    "annotation_pass",
    "source_annotator_id",
    "notes",
]

LABEL_VALUES = [
    "REQUEST",
    "INFORM_CONSTRAINT",
    "CONFIRM_ACCEPT",
    "CORRECT_CLARIFY",
    "SOCIAL",
]
PASS_VALUES = ["initial", "reannotation"]

COLUMN_WIDTHS = {
    "A": 25,
    "B": 18,
    "C": 9,
    "D": 60,
    "E": 60,
    "F": 24,
    "G": 14,
    "H": 15,
    "I": 20,
    "J": 40,
}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Create one dropdown-enabled Excel workbook per annotator with "
            "initial and reannotation sheets."
        )
    )
    parser.add_argument(
        "--initial-csv",
        default="data/processed/assignments_initial.csv",
        help="Path to initial assignment CSV.",
    )
    parser.add_argument(
        "--reannotation-csv",
        default="data/processed/assignments_reannotation.csv",
        help="Path to reannotation assignment CSV.",
    )
    parser.add_argument(
        "--out-dir",
        default="annotation/excel",
        help="Output directory for .xlsx workbooks.",
    )
    return parser.parse_args()


def read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        header = reader.fieldnames or []
        if header != CSV_FIELDNAMES:
            raise ValueError(
                f"Unexpected schema in {path}. "
                f"Expected {CSV_FIELDNAMES}, got {header}."
            )
        return [{k: row.get(k, "") for k in CSV_FIELDNAMES} for row in reader]


def group_by_annotator(rows: List[Dict[str, str]]) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    for row in rows:
        annotator = row.get("annotator_id", "").strip()
        if not annotator:
            raise ValueError("Found row with empty annotator_id.")
        grouped[annotator].append(row)
    return grouped


def _label_formula() -> str:
    return '"' + ",".join(LABEL_VALUES) + '"'


def _pass_formula() -> str:
    return '"' + ",".join(PASS_VALUES) + '"'


def build_source_owner_by_instance(initial_rows: List[Dict[str, str]]) -> Dict[str, str]:
    owners: Dict[str, str] = {}
    for row in initial_rows:
        instance_id = row.get("instance_id", "").strip()
        annotator_id = row.get("annotator_id", "").strip()
        if not instance_id:
            raise ValueError("Found initial row with empty instance_id.")
        if not annotator_id:
            raise ValueError("Found initial row with empty annotator_id.")
        prev = owners.get(instance_id)
        if prev is not None and prev != annotator_id:
            raise ValueError(
                f"Conflicting initial owner for {instance_id}: {prev} vs {annotator_id}"
            )
        owners[instance_id] = annotator_id
    return owners


def to_workbook_rows(
    rows: List[Dict[str, str]],
    source_owner_by_instance: Dict[str, str],
    mode: str,
) -> List[Dict[str, str]]:
    if mode not in {"initial", "reannotation"}:
        raise ValueError(f"Unexpected mode: {mode}")

    out: List[Dict[str, str]] = []
    for row in rows:
        item = {k: row.get(k, "") for k in WORKBOOK_FIELDNAMES if k != "source_annotator_id"}
        if mode == "initial":
            item["source_annotator_id"] = ""
        else:
            instance_id = row.get("instance_id", "").strip()
            source_owner = source_owner_by_instance.get(instance_id, "")
            if not source_owner:
                raise ValueError(
                    f"Could not resolve source annotator for reannotation instance_id={instance_id}"
                )
            item["source_annotator_id"] = source_owner
        out.append(item)
    return out


def apply_sheet_formatting(ws, n_rows: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{max(1, n_rows + 1)}"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, len(WORKBOOK_FIELDNAMES) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.protection = Protection(locked=False)

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    for r in range(2, n_rows + 2):
        for c in range(1, len(WORKBOOK_FIELDNAMES) + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    label_validation = DataValidation(
        type="list",
        formula1=_label_formula(),
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid label",
        error="Choose one of the five allowed labels.",
    )
    pass_validation = DataValidation(
        type="list",
        formula1=_pass_formula(),
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid pass",
        error="Choose initial or reannotation.",
    )
    ws.add_data_validation(label_validation)
    ws.add_data_validation(pass_validation)

    if n_rows > 0:
        label_validation.add(f"F2:F{n_rows + 1}")
        pass_validation.add(f"H2:H{n_rows + 1}")


def write_sheet(ws, rows: List[Dict[str, str]]) -> None:
    ws.append(WORKBOOK_FIELDNAMES)
    for row in rows:
        ws.append([row.get(k, "") for k in WORKBOOK_FIELDNAMES])
    apply_sheet_formatting(ws, len(rows))


def build_workbook(
    annotator: str,
    initial_rows: List[Dict[str, str]],
    reannotation_rows: List[Dict[str, str]],
    out_dir: Path,
) -> Path:
    wb = Workbook()
    ws_initial = wb.active
    ws_initial.title = "initial"
    write_sheet(ws_initial, initial_rows)

    ws_reannotation = wb.create_sheet("reannotation")
    write_sheet(ws_reannotation, reannotation_rows)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{annotator}_annotation.xlsx"
    wb.save(out_path)
    return out_path


def main() -> int:
    args = parse_args()
    initial_path = Path(args.initial_csv)
    reannotation_path = Path(args.reannotation_csv)
    out_dir = Path(args.out_dir)

    if not initial_path.exists():
        raise FileNotFoundError(f"Missing file: {initial_path}")
    if not reannotation_path.exists():
        raise FileNotFoundError(f"Missing file: {reannotation_path}")

    initial_rows = read_csv(initial_path)
    reannotation_rows = read_csv(reannotation_path)
    source_owner_by_instance = build_source_owner_by_instance(initial_rows)
    initial_by_annotator = group_by_annotator(initial_rows)
    reannotation_by_annotator = group_by_annotator(reannotation_rows)

    annotators = sorted(set(initial_by_annotator) | set(reannotation_by_annotator))
    if not annotators:
        raise ValueError("No annotators found in input CSV files.")

    print("Building workbooks:")
    for annotator in annotators:
        out_path = build_workbook(
            annotator=annotator,
            initial_rows=to_workbook_rows(
                initial_by_annotator.get(annotator, []),
                source_owner_by_instance=source_owner_by_instance,
                mode="initial",
            ),
            reannotation_rows=to_workbook_rows(
                reannotation_by_annotator.get(annotator, []),
                source_owner_by_instance=source_owner_by_instance,
                mode="reannotation",
            ),
            out_dir=out_dir,
        )
        print(
            f"- {annotator}: "
            f"{len(initial_by_annotator.get(annotator, []))} initial, "
            f"{len(reannotation_by_annotator.get(annotator, []))} reannotation -> "
            f"{out_path}"
        )

    print("Workbook generation complete")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
