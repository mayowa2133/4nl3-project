#!/usr/bin/env python3
"""Compile team annotations and compute overlap agreement metrics.

Outputs:
1. data/processed/annotations_labeled_long.csv
2. data/processed/annotations_overlap_pairs.csv
3. data/processed/agreement_summary.json
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

LABELS: List[str] = [
    "REQUEST",
    "INFORM_CONSTRAINT",
    "CONFIRM_ACCEPT",
    "CORRECT_CLARIFY",
    "SOCIAL",
]
ALLOWED_LABELS = set(LABELS)
REQUIRED_COLUMNS = [
    "instance_id",
    "dialogue_id",
    "turn_id",
    "system_context",
    "user_utterance",
    "label",
    "annotator_id",
    "annotation_pass",
    "notes",
]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compile annotation workbooks and compute Krippendorff alpha."
    )
    parser.add_argument(
        "--excel-dir",
        default="annotation/excel",
        help="Directory containing *_annotation.xlsx workbooks.",
    )
    parser.add_argument(
        "--long-out",
        default="data/processed/annotations_labeled_long.csv",
        help="Output path for long-form compiled annotations.",
    )
    parser.add_argument(
        "--overlap-out",
        default="data/processed/annotations_overlap_pairs.csv",
        help="Output path for overlap pair rows.",
    )
    parser.add_argument(
        "--summary-out",
        default="data/processed/agreement_summary.json",
        help="Output path for agreement summary JSON.",
    )
    return parser


def clean(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def workbook_paths(excel_dir: Path) -> List[Path]:
    paths = sorted(excel_dir.glob("*_annotation.xlsx"))
    return [p for p in paths if not p.name.startswith("~$")]


def required_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        if isinstance(raw, str):
            headers[raw.strip()] = col

    missing = [k for k in REQUIRED_COLUMNS if k not in headers]
    if missing:
        raise ValueError(f"Missing required columns in sheet '{ws.title}': {missing}")
    return headers


def sheet_rows(workbook_path: Path, sheet_name: str) -> List[Dict[str, str]]:
    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"{workbook_path}: missing sheet '{sheet_name}'")

    ws = wb[sheet_name]
    headers = required_headers(ws)
    source_col = headers.get("source_annotator_id")
    rows: List[Dict[str, str]] = []

    for row_idx in range(2, ws.max_row + 1):
        item: Dict[str, str] = {}
        for key in REQUIRED_COLUMNS:
            item[key] = clean(ws.cell(row=row_idx, column=headers[key]).value)
        item["source_annotator_id"] = (
            clean(ws.cell(row=row_idx, column=source_col).value) if source_col else ""
        )
        item["workbook"] = workbook_path.name

        # Ignore physically empty trailing rows if any.
        if not item["instance_id"] and not item["user_utterance"]:
            continue

        label = item["label"]
        if not label:
            raise ValueError(
                f"{workbook_path}:{sheet_name}: row {row_idx} has blank label."
            )
        if label not in ALLOWED_LABELS:
            raise ValueError(
                f"{workbook_path}:{sheet_name}: row {row_idx} has invalid label '{label}'."
            )

        rows.append(item)

    return rows


def compile_long_rows(excel_dir: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for path in workbook_paths(excel_dir):
        rows.extend(sheet_rows(path, "initial"))
        rows.extend(sheet_rows(path, "reannotation"))
    if not rows:
        raise ValueError(f"No workbook rows found in {excel_dir}")
    return rows


def write_csv(path: Path, fieldnames: Sequence[str], rows: Iterable[Dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def build_overlap_pairs(long_rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    initial_lookup: Dict[Tuple[str, str], Dict[str, str]] = {}
    overlap: List[Dict[str, str]] = []

    for row in long_rows:
        if row["annotation_pass"] == "initial":
            key = (row["instance_id"], row["annotator_id"])
            if key in initial_lookup:
                raise ValueError(
                    f"Duplicate initial annotation for instance_id={key[0]} annotator_id={key[1]}"
                )
            initial_lookup[key] = row

    for row in long_rows:
        if row["annotation_pass"] != "reannotation":
            continue

        source_annotator = row["source_annotator_id"]
        if not source_annotator:
            raise ValueError(
                "Reannotation row missing source_annotator_id for "
                f"instance_id={row['instance_id']}"
            )
        key = (row["instance_id"], source_annotator)
        if key not in initial_lookup:
            raise ValueError(
                "Could not match reannotation row to original row for "
                f"instance_id={row['instance_id']} source_annotator_id={source_annotator}"
            )
        original = initial_lookup[key]
        overlap.append(
            {
                "instance_id": row["instance_id"],
                "dialogue_id": row["dialogue_id"],
                "turn_id": row["turn_id"],
                "original_annotator_id": source_annotator,
                "reannotator_id": row["annotator_id"],
                "original_label": original["label"],
                "reannotation_label": row["label"],
            }
        )

    return overlap


def nominal_krippendorff_alpha_from_pairs(pairs: List[Tuple[str, str]]) -> float | None:
    """Nominal Krippendorff alpha for units with two labels each."""
    if not pairs:
        return None

    disagreements = sum(1 for a, b in pairs if a != b)
    observed_disagreement = disagreements / len(pairs)

    all_labels: List[str] = []
    for a, b in pairs:
        all_labels.extend([a, b])
    counts = Counter(all_labels)
    total = len(all_labels)
    if total <= 1:
        return None

    expected_agreement = sum((count / total) ** 2 for count in counts.values())
    expected_disagreement = 1.0 - expected_agreement

    if expected_disagreement == 0.0:
        return 1.0 if observed_disagreement == 0.0 else None

    return 1.0 - (observed_disagreement / expected_disagreement)


def confusion_matrix(
    pairs: List[Tuple[str, str]], labels: Sequence[str]
) -> Dict[str, Dict[str, int]]:
    matrix: Dict[str, Dict[str, int]] = {
        row_label: {col_label: 0 for col_label in labels} for row_label in labels
    }
    for original, reannotated in pairs:
        matrix[original][reannotated] += 1
    return matrix


def main() -> None:
    args = build_parser().parse_args()
    excel_dir = Path(args.excel_dir)
    long_out = Path(args.long_out)
    overlap_out = Path(args.overlap_out)
    summary_out = Path(args.summary_out)

    long_rows = compile_long_rows(excel_dir)
    overlap_rows = build_overlap_pairs(long_rows)

    long_fields = [
        "instance_id",
        "dialogue_id",
        "turn_id",
        "system_context",
        "user_utterance",
        "annotator_id",
        "annotation_pass",
        "source_annotator_id",
        "label",
        "notes",
        "workbook",
    ]
    write_csv(long_out, long_fields, long_rows)

    overlap_fields = [
        "instance_id",
        "dialogue_id",
        "turn_id",
        "original_annotator_id",
        "reannotator_id",
        "original_label",
        "reannotation_label",
    ]
    write_csv(overlap_out, overlap_fields, overlap_rows)

    pair_labels: List[Tuple[str, str]] = [
        (row["original_label"], row["reannotation_label"]) for row in overlap_rows
    ]
    matches = sum(1 for a, b in pair_labels if a == b)
    percent_agreement = (matches / len(pair_labels)) if pair_labels else 0.0
    alpha = nominal_krippendorff_alpha_from_pairs(pair_labels)

    initial_rows = sum(1 for row in long_rows if row["annotation_pass"] == "initial")
    reannotation_rows = sum(
        1 for row in long_rows if row["annotation_pass"] == "reannotation"
    )
    blank_label_count = sum(1 for row in long_rows if not row["label"])

    label_counts_original = Counter(row["original_label"] for row in overlap_rows)
    label_counts_reannotation = Counter(row["reannotation_label"] for row in overlap_rows)

    summary = {
        "initial_rows": initial_rows,
        "reannotation_rows": reannotation_rows,
        "overlap_rows": len(overlap_rows),
        "blank_label_count": blank_label_count,
        "percent_agreement": percent_agreement,
        "krippendorff_alpha_nominal": alpha,
        "label_counts_original": {
            label: label_counts_original.get(label, 0) for label in LABELS
        },
        "label_counts_reannotation": {
            label: label_counts_reannotation.get(label, 0) for label in LABELS
        },
        "confusion_matrix": confusion_matrix(pair_labels, LABELS),
    }

    summary_out.parent.mkdir(parents=True, exist_ok=True)
    with summary_out.open("w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"Compiled rows:      {len(long_rows)} -> {long_out}")
    print(f"Overlap pair rows:  {len(overlap_rows)} -> {overlap_out}")
    print(f"Agreement summary:  {summary_out}")
    print(f"Percent agreement:  {percent_agreement:.6f}")
    if alpha is None:
        print("Krippendorff alpha: undefined")
    else:
        print(f"Krippendorff alpha: {alpha:.6f}")


if __name__ == "__main__":
    main()
